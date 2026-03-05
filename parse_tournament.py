import pathlib

import requests
import pandas as pd
import re
import json
from bs4 import BeautifulSoup
from collections import defaultdict
from openpyxl import Workbook
from utils import get_mapper_from_nxt_data
import os
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

TOURNAMENT_URL = "https://proberserk.ru/tournament/74441f78-9220-434c-8f64-3f7a7f341a05"

DATA_FILEPATH = "data"
CARDS_JSON_FILEPATH = f"{DATA_FILEPATH}/cards.json"
CARDS_FILEPATH = f"{DATA_FILEPATH}/cards.xlsx"
COSTS_FILEPATH = f"{DATA_FILEPATH}/costs.xlsx"
ELEMENTS_FILEPATH = f"{DATA_FILEPATH}/elements.xlsx"
CARD_MATRIX_FILEPATH = f"{DATA_FILEPATH}/card_probability_matrix.xlsx"
ARCHETYPES_FILEPATH = f"{DATA_FILEPATH}/archetypes.xlsx"
DECKS_FILEPATH = f"{DATA_FILEPATH}/decks.xlsx"


def parse():
    """Функция парсит турнир и скачивает все деклисты в cards.json."""

    base_url = "https://proberserk.ru"
    response = requests.get(TOURNAMENT_URL)
    soup = BeautifulSoup(response.content, "html.parser")

    mapper = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for deck_tr in soup.find_all("tr"):
        a_tag = deck_tr.find("a", href=True)
        if not a_tag or not "/deck/" in a_tag["href"]:
            continue
        deck_url = base_url + a_tag["href"]
        deck_response = requests.get(deck_url)
        print(f"Получил колоду по ссылке {deck_url}")
        deck_soup = BeautifulSoup(deck_response.content, "html.parser")
        small_tag = deck_soup.find("small")
        player_name = small_tag.get_text().strip()
        h3_tag = deck_soup.find("h3")
        title = h3_tag.text.strip()
        title = title.replace(player_name, "")
        player_name = re.sub(r"\([^)]*\)", "", player_name).strip()
        print(f"Игрок {player_name} и колода {title}")
        for card_tr in deck_soup.find_all("tr"):
            card_td_tag = card_tr.find("td")
            card_a_tag = card_tr.find("a", href=True)
            if not card_td_tag or not card_a_tag:
                continue
            card_name = card_a_tag.get_text().strip()
            card_quantity = card_td_tag.get_text().strip()
            print(f"Карта {card_name} в количестве {card_quantity}")
            mapper[player_name][title][card_name] = int(card_quantity)

    with open(CARDS_JSON_FILEPATH, "w", encoding="utf-8") as file:
        file.write(json.dumps(mapper, ensure_ascii=False))
    print(f"Сохранил информацию по {len(mapper)} игрокам")


def create_cards_excel():
    """Функция обрабатывает cards.json и создает эксель со статистикой карт."""

    with open(CARDS_JSON_FILEPATH, "r", encoding="utf-8") as file:
        data = json.load(file)

    mapper = defaultdict(lambda: defaultdict(int))
    for player in data:
        for deck in data[player]:
            for card in data[player][deck]:
                mapper[card]["quantity"] += data[player][deck][card]
                mapper[card]["total_decks"] += 1
                for usage_copies in range(1, 4):
                    if data[player][deck][card] == usage_copies:
                        mapper[card][usage_copies] += 1

    wb = Workbook()
    ws = wb.active
    headers = ["Карта", "Кол-во", "Кол-во колод", "1 копии (кол-во колод)", "2 копии (кол-во колод)", "3 копии (кол-во колод)", "Средн. использование на колоду"]
    ws.append(headers)

    lines = []
    for card in mapper:
        quantity = mapper[card]["quantity"]
        total_decks = mapper[card]['total_decks']
        average_usage = round(mapper[card]["quantity"]/mapper[card]['total_decks'], 3)
        line = [card, quantity, total_decks, mapper[card][1], mapper[card][2], mapper[card][3], average_usage]
        print(f'{card}: {", ".join([str(_) for _ in line])}')
        lines.append(line)

    lines_sorted = sorted(lines, key=lambda x: x[1], reverse=True)
    df = pd.DataFrame(lines_sorted, columns=headers)
    with pd.ExcelWriter(CARDS_FILEPATH, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Карты", index=False)
        worksheet = writer.sheets["Карты"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 20, format_center)
        worksheet.set_column('D:AAA', 25, format_center)
        worksheet.set_column('G:AAA', 30, format_center)

    print(f"Создал эксель с картами с {len(lines)} строчками")


def create_probability_matrix_and_archetypes_excel():
    """Функция обрабатывает cards.json и создает два экселя: 1) матрица карт с вероятностями быть в одной колоде с друг другом; 2) статистика стейплов"""

    with open(CARDS_JSON_FILEPATH, "r", encoding="utf-8") as file:
        data = json.load(file)

    encountered_thrice_counts = defaultdict(int)
    card_deck_counts = defaultdict(int)
    pair_deck_counts = defaultdict(lambda: defaultdict(int))

    for player, decks in data.items():
        for deck_name, cards in decks.items():
            card_list = set(cards.keys())
            for card in card_list:
                if cards[card] >= 3:
                    encountered_thrice_counts[card] += 1
                card_deck_counts[card] += 1
                for other_card in card_list:
                    if card != other_card:
                        pair_deck_counts[card][other_card] += 1

    cards = sorted(card_deck_counts.keys())

    matrix = []
    for card_A in cards:
        row = []
        for card_B in cards:
            if card_A == card_B:
                row.append(float('nan'))
            else:
                cnt_A_and_B = pair_deck_counts[card_A][card_B]
                cnt_A = card_deck_counts[card_A]
                prob = cnt_A_and_B / cnt_A * 100 if cnt_A else 0
                row.append(round(prob, 3))
        matrix.append(row)

    card_matrix_df = pd.DataFrame(matrix, index=cards, columns=cards)
    with pd.ExcelWriter(CARD_MATRIX_FILEPATH, engine="xlsxwriter") as writer:
        card_matrix_df.to_excel(writer, sheet_name="Карты")
        worksheet = writer.sheets["Карты"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 20, format_center)
    print(f"Создал матрицу вероятностей карт с количеством {len(card_matrix_df)}")

    archetype_rows = []
    for card, encountered_thrice in sorted(encountered_thrice_counts.items(), key=lambda x: x[1], reverse=True):
        pairs = [(other, card_matrix_df.loc[card, other]) for other in card_matrix_df.columns if
                 other != card and card_matrix_df.loc[card, other] is not None]
        pairs_sorted = sorted(pairs, key=lambda x: x[1], reverse=True)
        row = {"Основная карта": card, "Встречаемость (3+ копии)": encountered_thrice}
        for idx, (other, prob) in enumerate(pairs_sorted, 1):
            if prob < 5:
                break
            row[f"Карта_{idx}"] = f"{other} ({prob}%)"
        archetype_rows.append(row)
    archetypes_df = pd.DataFrame(archetype_rows)

    encountered_cards = []
    archetype_without_copies_rows = []
    for card, encountered_thrice in sorted(encountered_thrice_counts.items(), key=lambda x: x[1], reverse=True):
        pairs = [(other, card_matrix_df.loc[card, other]) for other in card_matrix_df.columns if
                 other != card and card_matrix_df.loc[card, other] is not None]
        pairs_sorted = sorted(pairs, key=lambda x: x[1], reverse=True)
        row = {"Основная карта": card, "Встречаемость (3+ копии)": encountered_thrice}
        for idx, (other, prob) in enumerate(pairs_sorted, 1):
            if prob < 40:
                break
            encountered_cards.append(other)
            row[f"Карта_{idx}"] = f"{other} ({prob}%)"
        if card not in encountered_cards:
            archetype_without_copies_rows.append(row)
    archetype_without_copies_df = pd.DataFrame(archetype_without_copies_rows)

    with pd.ExcelWriter(ARCHETYPES_FILEPATH, engine="xlsxwriter") as writer:
        archetypes_df.to_excel(writer, sheet_name="Архетипы", index=False)
        worksheet = writer.sheets["Архетипы"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:A', 25, format_center)
        worksheet.set_column('B:B', 25, format_center)
        worksheet.set_column('C:AAA', 30, format_center)
        print(f"Создал файл с архетипами с количеством {len(archetypes_df)}")

        archetype_without_copies_df.to_excel(writer, sheet_name="Архетипы без копий", index=False)
        worksheet = writer.sheets["Архетипы без копий"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:A', 25, format_center)
        worksheet.set_column('B:B', 25, format_center)
        worksheet.set_column('C:AAA', 30, format_center)
        print(f"Создал файл с архетипами без копий с количеством {len(archetype_without_copies_df)}")


def create_decks_excel():
    """Функция обрабатывает cards.json и создает эксель со статистикой архетипов (нужно поменять archetypes.txt - через запятую указать карты в архетипе, последняя строчка это название архетипа)."""

    filename = CARDS_JSON_FILEPATH
    mapper = defaultdict(int)
    unknown_decks = []
    not_found_decklists = []

    total_lineaps = 0
    mapper_for_percentage = defaultdict(int)

    with open("data/archetypes.txt", "r", encoding="utf-8") as file:
        file_data = file.readlines()
        archetypes = {}
        for line in file_data:
            line = line.replace("\n", "").strip()
            line_arr = line.split(", ")
            archetypes[tuple(line_arr[:-1])] = line_arr[-1]

    for path, folders, files in os.walk(DATA_FILEPATH):
        for filename in files:
            if not filename == "cards.json" or "final" in path:
                continue
            filename = os.path.join(path, filename)

            print(f"Открываю файл для обработки колод {filename}")

            with open(filename, "r", encoding="utf-8") as file:
                data = json.load(file)

            for player, decks in data.items():
                total_lineaps += 1
                for deck_name, cards in decks.items():
                    card_list = set(cards.keys())
                    for archetype in archetypes:
                        for card in archetype:
                            if card not in card_list:
                                break
                        else:
                            mapper[archetypes[archetype]] += 1
                            mapper_for_percentage[archetypes[archetype]] += 1
                            print(f"Добавил деклист к {archetypes[archetype]}")
                            break
                    else:
                        print(f"Не удалось подобрать архетип для колоды: {card_list}")
                        unknown_decks.append(card_list)
                        not_found_decklists.append(str(card_list))

    lines = [[archetype, mapper[archetype], round(mapper_for_percentage[archetype] / total_lineaps * 100, 2)] for archetype in mapper]
    lines.append(["Авторская сборка", len(unknown_decks), round(len(unknown_decks) / total_lineaps * 100, 2)])
    df = pd.DataFrame(sorted(lines, key=lambda x: x[1], reverse=True), columns=["Колода", "Кол-во", "Наличие в лайнапе, %"])
    with pd.ExcelWriter("data/total_decks.xlsx", engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Колоды", index=False)
        worksheet = writer.sheets["Колоды"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:A', 25, format_center)
        worksheet.set_column('B:B', 25, format_center)
        worksheet.set_column('C:AAA', 30, format_center)
        print(f"Создал файл с колодами в количестве {len(df)}")

        not_found_decklists_df = pd.DataFrame(not_found_decklists, columns=["Неизвестные колоды"])
        not_found_decklists_df.to_excel(writer, sheet_name="Неизвестные колоды", index=False)
        worksheet = writer.sheets["Неизвестные колоды"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 80, format_center)


def create_costs_excel():
    """Функция обрабатывает cards.json и создает эксель со статистикой стоимостей."""

    nxt_mapper = get_mapper_from_nxt_data()

    with open(CARDS_JSON_FILEPATH, "r", encoding="utf-8") as file:
        data = json.load(file)

    basic_mapper = defaultdict(lambda: defaultdict(int))
    elite_mapper = defaultdict(lambda: defaultdict(int))
    for player in data:
        for deck in data[player]:
            for card in data[player][deck]:
                cost = nxt_mapper[card]["cost"]
                elite = nxt_mapper[card]["elite"]
                if elite:
                    mapper = elite_mapper
                else:
                    mapper = basic_mapper
                mapper[cost]["quantity"] += data[player][deck][card]
                mapper[cost]["total_quantity"] += 1
                mapper[cost][card] += data[player][deck][card]
                for usage_copies in range(1, 4):
                    if data[player][deck][card] == usage_copies:
                        mapper[cost][usage_copies] += 1

    short_headers = ["Стоимость", "Кол-во", "Кол-во колод"]
    headers = short_headers + ["1 копии (кол-во колод)", "2 копии (кол-во колод)", "3 копии (кол-во колод)"]
    top_cards_quantity = 5
    for _ in range(1, top_cards_quantity+1):
        short_headers.append(f"Карта_{_}")
        headers.append(f"Карта_{_}")

    basic_short_lines = []
    basic_lines = []
    for cost in basic_mapper:
        quantity = basic_mapper[cost].pop("quantity")
        total_quantity = basic_mapper[cost].pop("total_quantity")
        cards = list(map(lambda x: f"{x[0]} ({x[1]})", sorted(filter(lambda x: isinstance(x[0], str), basic_mapper[cost].items()), key=lambda x: x[1], reverse=True)[:top_cards_quantity]))
        while len(cards) < top_cards_quantity:
            cards.append("")
        short_line = [cost, quantity, total_quantity]
        line = short_line + [basic_mapper[cost][1], basic_mapper[cost][2], basic_mapper[cost][3]]
        print(f'{cost}: {", ".join([str(_) for _ in line])}')
        basic_short_lines.append(short_line + cards)
        basic_lines.append(line + cards)

    elite_short_lines = []
    elite_lines = []
    for cost in elite_mapper:
        quantity = elite_mapper[cost].pop("quantity")
        total_quantity = elite_mapper[cost].pop("total_quantity")
        cards = list(map(lambda x: f"{x[0]} ({x[1]})", sorted(filter(lambda x: isinstance(x[0], str), elite_mapper[cost].items()), key=lambda x: x[1], reverse=True)[:top_cards_quantity]))
        while len(cards) < top_cards_quantity:
            cards.append("")
        short_line = [cost, quantity, total_quantity]
        line = short_line + [elite_mapper[cost][1], elite_mapper[cost][2], elite_mapper[cost][3]]
        print(f'{cost}: {", ".join([str(_) for _ in line])}')
        elite_short_lines.append(short_line + cards)
        elite_lines.append(line + cards)

    basic_short_lines_sorted = sorted(basic_short_lines, key=lambda x: x[0])
    basic_lines_sorted = sorted(basic_lines, key=lambda x: x[0])
    elite_short_lines_sorted = sorted(elite_short_lines, key=lambda x: x[0])
    elite_lines_sorted = sorted(elite_lines, key=lambda x: x[0])

    basic_short_df = pd.DataFrame(basic_short_lines_sorted, columns=short_headers)
    basic_df = pd.DataFrame(basic_lines_sorted, columns=headers)
    elite_short_df = pd.DataFrame(elite_short_lines_sorted, columns=short_headers)
    elite_df = pd.DataFrame(elite_lines_sorted, columns=headers)
    with pd.ExcelWriter(COSTS_FILEPATH, engine="xlsxwriter") as writer:

        basic_short_df.to_excel(writer, sheet_name="Рядовые (мин)", index=False)
        worksheet = writer.sheets["Рядовые (мин)"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 15, format_center)
        worksheet.set_column('D:AAA', 25, format_center)
        worksheet.set_column('G:AAA', 30, format_center)

        basic_df.to_excel(writer, sheet_name="Рядовые", index=False)
        worksheet = writer.sheets["Рядовые"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 15, format_center)
        worksheet.set_column('D:AAA', 25, format_center)
        worksheet.set_column('G:AAA', 30, format_center)

        elite_short_df.to_excel(writer, sheet_name="Элитные (мин)", index=False)
        worksheet = writer.sheets["Элитные (мин)"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 15, format_center)
        worksheet.set_column('D:AAA', 25, format_center)
        worksheet.set_column('G:AAA', 30, format_center)

        elite_df.to_excel(writer, sheet_name="Элитные", index=False)
        worksheet = writer.sheets["Элитные"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 15, format_center)
        worksheet.set_column('D:AAA', 25, format_center)
        worksheet.set_column('G:AAA', 30, format_center)

    print(f"Создал эксель с рядовыми стоимостями на {len(basic_df)} строчками")
    print(f"Создал эксель с элитными стоимостями на {len(elite_df)} строчками")


def create_elements_excel():
    """Функция обрабатывает cards.json и создает эксель со статистикой стихий."""

    nxt_mapper = get_mapper_from_nxt_data()

    with open(CARDS_JSON_FILEPATH, "r", encoding="utf-8") as file:
        data = json.load(file)

    quantity_field_name = 100
    total_quantity_field_name = 200
    mapper = defaultdict(lambda: defaultdict(int))
    for player in data:
        for deck in data[player]:
            for card in data[player][deck]:
                color = nxt_mapper[card]["color"]
                mapper[color][quantity_field_name] += data[player][deck][card]
                mapper[color][total_quantity_field_name] += 1
                mapper[color][card] += data[player][deck][card]
                for usage_copies in range(1, 4):
                    if data[player][deck][card] == usage_copies:
                        mapper[color][usage_copies] += 1

    wb = Workbook()
    ws = wb.active
    short_headers = ["Стихия", "Кол-во", "Кол-во колод"]
    headers = short_headers + ["1 копии (кол-во колод)", "2 копии (кол-во колод)", "3 копии (кол-во колод)"]
    top_cards_quantity = 5
    for _ in range(1, top_cards_quantity+1):
        short_headers.append(f"Карта_{_}")
        headers.append(f"Карта_{_}")

    ws.append(headers)

    elements_mapper = {
        1: "Степи",
        2: "Горы",
        4: "Леса",
        8: "Болота",
        16: "Тьма",
        32: "Нейтралы",
    }

    lines = []
    for color in elements_mapper:
        mapped_color = elements_mapper[color]
        quantity = mapper[color][quantity_field_name]
        total_quantity = mapper[color][total_quantity_field_name]
        cards = list(map(lambda x: f"{x[0]} ({x[1]})", sorted(filter(lambda x: isinstance(x[0], str), mapper[color].items()), key=lambda x: x[1], reverse=True)[:top_cards_quantity]))
        while len(cards) < top_cards_quantity:
            cards.append("")
        short_line = [mapped_color, quantity, total_quantity]
        line = short_line + [mapper[color][1], mapper[color][2], mapper[color][3]]
        print(f'{mapped_color}: {", ".join([str(_) for _ in line])}')
        lines.append(line + cards)

    df = pd.DataFrame(lines, columns=headers)
    with pd.ExcelWriter(ELEMENTS_FILEPATH, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Стихии", index=False)
        worksheet = writer.sheets["Стихии"]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 15, format_center)
        worksheet.set_column('D:AAA', 25, format_center)
        worksheet.set_column('G:AAA', 30, format_center)

        worksheet = writer.book.add_worksheet("Стихии (мин)")
        format_center = writer.book.add_format({'align': 'center', 'valign': 'vcenter'})

        row = 0
        col = 0
        count = 0
        elements_list = list(elements_mapper.items())

        for idx, (color, mapped_color) in enumerate(elements_list):
            quantity = mapper[color][quantity_field_name]
            total_quantity = mapper[color][total_quantity_field_name]
            cards = list(map(lambda x: f"{x[0]} ({x[1]})",
                             sorted(filter(lambda x: isinstance(x[0], str), mapper[color].items()),
                                    key=lambda x: x[1], reverse=True)[:top_cards_quantity]))
            while len(cards) < top_cards_quantity:
                cards.append("")

            worksheet.write(row, col, short_headers[0], format_center)
            worksheet.write(row, col + 1, mapped_color, format_center)
            row += 1

            worksheet.write(row, col, short_headers[1], format_center)
            worksheet.write(row, col + 1, quantity, format_center)
            row += 1

            worksheet.write(row, col, short_headers[2], format_center)
            worksheet.write(row, col + 1, total_quantity, format_center)
            row += 1

            for card_idx, card in enumerate(cards, start=0):
                worksheet.write(row, col, f"Карта_{card_idx + 1}", format_center)
                worksheet.write(row, col + 1, card, format_center)
                row += 1

            row += 1
            count += 1

            if count == 3:
                col += 3
                count = 0
                row = 0

        worksheet.set_column('A:A', 15, format_center)
        worksheet.set_column('B:B', 30, format_center)
        worksheet.set_column('C:C', 10, format_center)
        worksheet.set_column('D:D', 15, format_center)
        worksheet.set_column('E:E', 30, format_center)
        for ind in range(100):
            worksheet.set_row(ind, 20, format_center)

    print(f"Создал эксель со стихиями с {len(lines)} строчками")


def create_common_excel():
    """Функция обрабатывает cards.json и создает эксель со статистикой рядовых и элитных карт."""

    sheet_name = "Рядовые"
    output_name = "total_costs.xlsx"
    file_name = "costs.xlsx"
    headers = ["Стоимость", "Кол-во", "Кол-во колод", "1 копии (кол-во колод)", "2 копии (кол-во колод)", "3 копии (кол-во колод)", "Средн. использование на колоду"]

    mapper = defaultdict(lambda: defaultdict(int))

    for path, folders, files in os.walk(DATA_FILEPATH):
        for filename in files:
            if not filename == file_name:
                continue

            try:
                print(f"Приступил к чтению файла {pathlib.Path(path, filename)}")

                df = pd.read_excel(pathlib.Path(path, filename), sheet_name=sheet_name)
                for idx, row in df.iterrows():
                    mapper[row[headers[0]]]["quantity"] += row[headers[1]]
                    mapper[row[headers[0]]]["total_decks"] += row[headers[2]]
                    mapper[row[headers[0]]][1] += row[headers[3]]
                    mapper[row[headers[0]]][2] += row[headers[4]]
                    mapper[row[headers[0]]][3] += row[headers[5]]

            except Exception as e:
                print(f"Произошла ошибка при создании общего эксель файла: {e}")
                continue

    wb = Workbook()
    ws = wb.active
    ws.append(headers)

    lines = []
    for card in mapper:
        quantity = mapper[card]["quantity"]
        total_decks = mapper[card]["total_decks"]
        average_usage = round(mapper[card]["quantity"] / mapper[card]["total_decks"], 3)
        line = [card, quantity, total_decks, mapper[card][1], mapper[card][2], mapper[card][3], average_usage]
        print(f'{card}: {", ".join([str(_) for _ in line])}')
        lines.append(line)

    # lines_sorted = sorted(lines, key=lambda x: x[1], reverse=True)
    df = pd.DataFrame(lines, columns=headers)
    with pd.ExcelWriter(DECKS_FILEPATH, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        format_center = writer.book.add_format({'align': 'center'})
        worksheet.set_column('A:AAA', 20, format_center)
        worksheet.set_column('D:AAA', 25, format_center)
        worksheet.set_column('G:AAA', 30, format_center)

    print(f"Создал общий эксель с {len(lines)} строчками")


def create_diagram():
    """Функция составляет диаграмму по деклистам, использовалась в Кристалле."""

    elements_mapper = {
        "Степи": "#FFC400",
        "Горы": "#0055FF",
        "Леса": "#00FF4D",
        "Болота": "#277F0C",
        "Тьма": "#202020",
        "Нейтралы": "#FF0000",
    }

    x_field_name = "Колода"
    y_field_name = "Кол-во"
    sheet_name = "Колоды"

    file_name = "decks.xlsx"

    plot_name = "Динамика популярности колод"
    x_name = "Тур"
    y_name = "Место"

    tours = range(1, 8)
    tour_dfs = {}
    tour_top = {}

    for t in tours:
        path = os.path.join("data", str(t), file_name)
        df = pd.read_excel(path, sheet_name=sheet_name)
        df = df.rename(columns={
            x_field_name: "x",
            y_field_name: "y"
        })
        tour_dfs[t] = df
        tour_top[t] = df.groupby("x")["y"].sum().sort_values(ascending=False).index

    all_top_cards = sorted(set().union(*tour_top.values()))

    pivot_data = []
    for t in tours:
        df_t = tour_dfs[t].copy()
        df_t["tour"] = t
        df_t = df_t[df_t["x"].isin(all_top_cards)]
        pivot_data.append(df_t)

    pivot_df = pd.concat(pivot_data, ignore_index=True)
    pivot = pivot_df.pivot_table(
        index="tour",
        columns="x",
        values="y",
        aggfunc="sum"
    )

    pivot_ranked = pivot.rank(axis=1, ascending=False, method='first').fillna(30)
    cm = plt.get_cmap('tab20')
    colors = cm(np.linspace(0, 1, len(pivot_ranked.columns)))

    fig, ax = plt.subplots(figsize=(14, 8))

    for val, color in zip(pivot_ranked.columns, colors):
        y = pivot_ranked[val].values
        x = pivot_ranked.index.values
        ax.plot(x, y, marker="o", color=color, linewidth=2)

        ax.annotate(
            val,
            xy=(x[-1], y[-1]),
            xytext=(5, 0),
            textcoords="offset points",
            va="center",
            ha="left",
            color=color,
            fontsize=8,
            fontweight="bold"
        )

    ax.set_title(plot_name, fontsize=12)
    ax.set_xlabel(x_name, fontsize=12)
    ax.set_ylabel(y_name, fontsize=12)
    ax.set_xticks(list(tours))
    ax.set_yticks(range(1, 21))
    ax.set_ylim(20.5, -0.5)
    ax.grid(True, alpha=0.3)
    ax.legend().set_visible(False)
    plt.tight_layout()
    plt.savefig("plot.png", dpi=300)


if __name__ == "__main__":
    parse()
    create_cards_excel()
    create_probability_matrix_and_archetypes_excel()
    # create_decks_excel() нужно поменять archetypes.txt и потом запускать
    create_costs_excel()
    create_elements_excel()
    create_common_excel()
    # create_diagram() нужно поменять под конкретные ваши нужды
